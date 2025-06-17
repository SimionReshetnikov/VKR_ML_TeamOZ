import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import cv2
import numpy as np
import tempfile

class ReportGenerator:
    def __init__(self, report_dir="reports"):
        self.report_dir = report_dir
        os.makedirs(self.report_dir, exist_ok=True)

    def save_report(
        self,
        results_per_frame,
        video_name="video",
        frame_images=None,
        extra_metrics=None,
    ):
        """
        results_per_frame: список словарей по кадрам
            [
                {
                    'frame_idx': int,
                    'objects': [
                        {
                            'bbox': [x1, y1, x2, y2],
                            'object_class': str,
                            'object_conf': float,
                            'defect_class': str,
                            'defect_conf': float
                        },
                        ...
                    ]
                },
                ...
            ]
        frame_images: список numpy.ndarray (BGR) — скриншоты соответствующих кадров
        extra_metrics: dict — необязательные метрики (например, % кадров с дефектами)
        """

        report_name = f"{video_name}_defect_report.xlsx"
        report_path = os.path.join(self.report_dir, report_name)

        rows = []
        img_files = []

        for idx, frame in enumerate(results_per_frame):
            frame_idx = frame.get("frame_idx", idx)
            objects = frame.get("objects", [])
            img = None
            if frame_images is not None and idx < len(frame_images):
                img = frame_images[idx]
            img_file = None
            if img is not None:
                # Сохраним временный файл для вставки в Excel
                fd, img_file = tempfile.mkstemp(suffix=".jpg")
                os.close(fd)
                cv2.imwrite(img_file, img)
                img_files.append(img_file)

            # --- ФИЛЬТРАЦИЯ: только дефекты, только defect_conf >= 0.7 ---
            defect_objects = [
                obj for obj in objects
                if "good" not in obj["defect_class"] and obj["defect_conf"] >= 0.7
            ]

            if not defect_objects:
                continue  # Если дефектов нет — не добавляем строку

            for obj in defect_objects:
                rows.append(
                    {
                        "Кадр": frame_idx,
                        "Класс объекта": obj["object_class"],
                        "Достоверность объекта": f'{obj["object_conf"]:.2f}',
                        "Класс дефекта": obj["defect_class"],
                        "Достоверность дефекта": f'{obj["defect_conf"]:.2f}',
                        "Скриншот": img_file if img is not None else "",
                    }
                )

        # DataFrame для отчёта
        if rows:
            df = pd.DataFrame(rows)
        else:
            # Если дефектов нет — создаём пустой DataFrame с нужными колонками
            df = pd.DataFrame(
                columns=[
                    "Кадр",
                    "Класс объекта",
                    "Достоверность объекта",
                    "Класс дефекта",
                    "Достоверность дефекта",
                    "Скриншот",
                ]
            )

        # Сохраняем в Excel
        df.to_excel(report_path, index=False, engine="openpyxl")

        # Вставляем изображения в Excel (openpyxl)
        wb = load_workbook(report_path)
        ws = wb.active

        img_col_idx = df.columns.get_loc("Скриншот") + 1
        img_col_letter = get_column_letter(img_col_idx)

        # Для красоты: ширина колонки под картинки
        ws.column_dimensions[img_col_letter].width = 24

        # Вставляем картинки (только по первой строке каждого кадра)
        prev_frame = None
        for i, row in enumerate(rows, start=2):  # openpyxl: первая строка — заголовки
            img_path = row["Скриншот"]
            frame = row["Кадр"]
            if img_path and (frame != prev_frame or i == 2):
                try:
                    img = XLImage(img_path)
                    img.height = 120
                    img.width = 180
                    cell = f"{img_col_letter}{i}"
                    ws.add_image(img, cell)
                except Exception as e:
                    print(f"Ошибка вставки изображения: {e}")
            prev_frame = frame

        # Метрики — отдельный лист
        if extra_metrics:
            ws2 = wb.create_sheet("Метрики")
            for idx, (k, v) in enumerate(extra_metrics.items(), start=1):
                ws2[f"A{idx}"] = k
                ws2[f"B{idx}"] = v

        wb.save(report_path)

        # Удаляем временные изображения
        for f in img_files:
            try:
                os.remove(f)
            except Exception:
                pass

        print(f"Отчёт сохранён: {report_path}")
        return report_path